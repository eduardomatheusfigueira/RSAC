#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Pipeline de Coleta e Raspagem de Metadados da BDTD (Biblioteca Digital Brasileira de Teses e Dissertações)
Autor: Eduardo Matheus Figueira
Descrição: Pipeline automatizado que consulta a BDTD utilizando a API bdtd-scraper,
             realiza raspagem (scraping) de metadados em páginas web utilizando requests.Session (Keep-Alive),
             persiste os dados limpos no banco SQLite (com transações otimizadas e UPSERT) e exporta os resultados.
"""

import os  # Módulo para interação com o sistema operacional (manipulação de caminhos e arquivos)
import sys  # Módulo para parâmetros e funções específicas do sistema (acesso a stdout, argumentos)
import time  # Módulo para controle de tempo, pausas (sleep) e medição de duração de execução
import sqlite3  # Biblioteca nativa para gerenciamento do banco de dados relacional SQLite
import logging  # Módulo para geração e emissão de mensagens de log no console/arquivo
import argparse  # Módulo para criação de interfaces de linha de comando (argumentos via terminal)
import unicodedata  # Módulo para normalização Unicode (remoção de acentos e caracteres especiais)
import requests  # Biblioteca HTTP requests com suporte a requisições persistentes (Session Keep-Alive)
import re  # Módulo para manipulação e busca de padrões com Expressões Regulares (Regex)
import pandas as pd  # Biblioteca Pandas para estruturação de dados em DataFrames e exportação
from bs4 import BeautifulSoup  # Biblioteca BeautifulSoup para parsing e extração de elementos em HTML
import bdtd_scraper.api  # Módulo de integração com a API pública do motor de busca da BDTD
from typing import List, Dict, Optional, Set, Tuple, Any, Union  # Importação de anotações de tipo para Type Hints

# Objeto de log do módulo
logger: logging.Logger = logging.getLogger(__name__)


def setup_logging(log_file_path: Optional[str] = None) -> None:
    """
    Configura o sistema de log centralizado evitando duplicação de handlers se o script for chamado/importado múltiplas vezes.
    """
    root_logger = logging.getLogger()  # Pega o registrador raiz (root logger)
    root_logger.setLevel(logging.INFO)  # Define o nível mínimo de log para INFO
    
    # Limpa handlers pré-existentes para evitar mensagens duplicadas no console/arquivo
    if root_logger.hasHandlers():
        root_logger.handlers.clear()
        
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')  # Formato: Data/Hora - Nível - Mensagem
    
    # Handler de saída no console (sys.stdout)
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    root_logger.addHandler(stream_handler)
    
    # Handler para arquivo de log (se fornecido um caminho válido)
    if log_file_path:
        try:
            log_dir = os.path.dirname(log_file_path) if os.path.dirname(log_file_path) else "."
            os.makedirs(log_dir, exist_ok=True)
            open_path = log_file_path
            if sys.platform.startswith('win') and len(os.path.abspath(open_path)) >= 250 and not os.path.abspath(open_path).startswith('\\\\?\\'):
                open_path = '\\\\?\\' + os.path.abspath(open_path)
            file_handler = logging.FileHandler(open_path, mode='w', encoding='utf-8')
            file_handler.setFormatter(formatter)
            root_logger.addHandler(file_handler)
            logger.info(f"Log de execução sendo gravado em: {log_file_path}")
        except Exception as e:
            logger.warning(f"Não foi possível iniciar o arquivo de log para gravação: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# EXPRESSÕES REGULARES PRÉ-COMPILADAS (OTIMIZAÇÃO DE PERFORMANCE DE CPU)
# Compilar no nível do módulo evita recompilações desnecessárias a cada chamada
# ──────────────────────────────────────────────────────────────────────────────
RE_PARTS_SPLIT: re.Pattern = re.compile(r'[;\n]+')  # Divisor de textos por ponto e vírgula ou quebras de linha
RE_YEAR_TRAILING: re.Pattern = re.compile(r'(?:,\s*|\s*)\d{4}-?\d*$', re.IGNORECASE)  # Remove anos tipo ", 1980-" ou " 1980-2020" no final
RE_ORIENTADOR_LABEL: re.Pattern = re.compile(r'^(?:orientador|orientadora|orientadores|orientador\(a\))\s*:\s*(.+)$', re.IGNORECASE)  # Identifica "Orientador: Nome"

# ──────────────────────────────────────────────────────────────────────────────
# RESTRIÇÃO DA API DA BDTD: O servidor VuFind da BDTD utiliza um Firewall de Aplicação Web (WAF)
# que bloqueia requisições HTTP contendo 3 ou mais parâmetros de filtro `filter[]`,
# retornando um erro HTTP 429 (Too Many Requests). Esta é uma regra do lado do servidor.
#
# Para contornar isso, limitamos a no máximo 2 filtros na API por requisição.
# Filtros de idioma (ex: ~language:por) são separados e aplicados localmente nos registros.
# ──────────────────────────────────────────────────────────────────────────────
BDTD_MAX_API_FILTERS: int = 2  # Limite máximo de filtros permitidos diretamente na chamada da API


def sanitize_bdtd_filters(raw_filters: Optional[List[str]]) -> Tuple[List[str], Set[str]]:
    """
    Separa uma lista de strings de filtro Solr da BDTD em:
      - api_filters: filtros seguros para envio como parâmetros filter[] na API (máximo BDTD_MAX_API_FILTERS)
      - allowed_langs: conjunto de códigos de idioma para filtragem pós-processamento realizada localmente

    Quaisquer filtros que não sejam de idioma além do limite BDTD_MAX_API_FILTERS são descartados com um aviso.

    Retorna:
        Tuple[List[str], Set[str]]: (filtros_api, idiomas_permitidos)
    """
    api_filters: List[str] = []  # Lista que guardará os filtros formatados para a requisição da API
    allowed_languages: Set[str] = set()  # Conjunto de idiomas válidos para filtragem local rápida

    # Itera sobre os filtros brutos recebidos
    for f in (raw_filters or []):
        if not f or not isinstance(f, str):  # Valida se o item existe e é uma string
            continue  # Pula itens nulos ou inválidos
        f_str: str = f.strip()  # Remove espaços extras nas extremidades
        if not f_str:  # Se a string estiver vazia após o strip
            continue  # Pula para o próximo item

        # Trata filtros com acoplamento por til "~" (ex: ~language:por ou language:por)
        stripped: str = f_str.lstrip('~')  # Remove o caractere til inicial se presente
        if stripped.startswith('language:'):  # Verifica se é uma chave de idioma
            lang_code: str = stripped.split(':', 1)[1].strip()  # Extrai o código do idioma (ex: por, eng)
            if lang_code:  # Se o código do idioma não for vazio
                allowed_languages.add(lang_code)  # Adiciona ao conjunto de idiomas permitidos
        else:
            # Garante que seja um par chave:valor válido com valor não vazio
            if ':' in f_str:  # Verifica se existe o separador de chave/valor
                field, val = f_str.split(':', 1)  # Divide no primeiro caractere de dois pontos
                if field.strip() and val.strip():  # Verifica se o campo e o valor contêm texto
                    api_filters.append(f_str)  # Inclui na lista de filtros da API
            else:
                logger.warning(f"Ignorando filtro BDTD inválido/vazio: '{f_str}'")  # Registra alerta para filtro malformado

    # Aplica o limite máximo exigido pelo WAF da BDTD
    if len(api_filters) > BDTD_MAX_API_FILTERS:  # Caso a quantidade de filtros supere o limite do WAF
        dropped: List[str] = api_filters[BDTD_MAX_API_FILTERS:]  # Identifica os filtros excedentes descartados
        api_filters = api_filters[:BDTD_MAX_API_FILTERS]  # Trunca a lista mantendo apenas os primeiros permitidos
        logger.warning(  # Emite mensagem de aviso no log informando o ajuste
            f"BDTD WAF limit: Too many API filters ({len(api_filters) + len(dropped)}). "
            f"Keeping first {BDTD_MAX_API_FILTERS}: {api_filters}. "
            f"Dropped: {dropped}"
        )

    return api_filters, allowed_languages  # Retorna a tupla contendo a lista tratada e o conjunto de idiomas


def clean_creator_name(creator_str: Optional[str]) -> str:
    """
    Limpa nomes de autores/criadores removendo anos de nascimento/falecimento no final
    utilizando a expressão regular otimizada pré-compilada RE_YEAR_TRAILING.
    """
    if not creator_str:  # Se a string do criador for nula ou vazia
        return ""  # Retorna string vazia
    parts: List[str] = RE_PARTS_SPLIT.split(creator_str)  # Divide por ponto e vírgula ou quebras de linha
    cleaned_parts: List[str] = []  # Lista para armazenar as partes limpas de cada autor
    for part in parts:  # Percorre cada parte dividida
        part = part.strip()  # Remove espaços extras nas extremidades
        if not part:  # Se a parte estiver vazia
            continue  # Pula para a próxima
        part = RE_YEAR_TRAILING.sub('', part)  # Aplica a regex pré-compilada para remover anos no final
        part = part.strip(',;.- ')  # Remove pontuações desnecessárias sobressalentes
        if part:  # Se a parte limpa contiver texto válido
            cleaned_parts.append(part)  # Adiciona à lista final de nomes limpos
    return "; ".join(cleaned_parts)  # Junta os nomes limpos utilizando ponto e vírgula como separador


def clean_advisor_name(advisor_str: Optional[str]) -> str:
    """
    Limpa o nome do orientador removendo links de currículo Lattes, URLs ORCID,
    departamentos institucionais e datas de nascimento/morte.
    """
    if not advisor_str or advisor_str.lower() == "não informado":  # Se o orientador for nulo ou "não informado"
        return "Não Informado"  # Retorna o texto padrão unificado
    
    parts: List[str] = RE_PARTS_SPLIT.split(advisor_str)  # Divide a string utilizando a regex pré-compilada
    cleaned_parts: List[str] = []  # Lista para guardar os nomes de orientadores validados
    
    for part in parts:  # Itera sobre cada trecho do nome do orientador
        part = part.strip()  # Remove espaços em branco ao redor
        if not part:  # Se estiver vazio
            continue  # Pula o item
            
        # 1. Filtra e descarta links e URLs (Lattes, ORCID, etc.)
        if part.startswith('http://') or part.startswith('https://') or 'lattes.cnpq.br' in part or 'orcid.org' in part:
            continue  # Descarta o trecho que contém link de internet
            
        # 2. Filtra e descarta informações institucionais errôneas no campo de orientador
        lower_part: str = part.lower()  # Converte o trecho para minúsculas para comparação
        is_institution: bool = any(word in lower_part for word in [  # Verifica palavras institucionais comuns
            'universidade', 'faculdade', 'instituto', 'programa de', 'departamento',
            'campus', 'reitor', 'pró-reitoria', 'coordenador', 'membro da banca'
        ])
        if is_institution:  # Se for identificado como nome de instituição
            continue  # Descarta o trecho
            
        # 3. Limpa anos de nascimento/morte no final do nome utilizando a regex pré-compilada
        part = RE_YEAR_TRAILING.sub('', part)
        
        part = part.strip(',;.- ')  # Remove pontuações desnecessárias das pontas
        if part:  # Se o nome restar válido
            cleaned_parts.append(part)  # Adiciona ao grupo de orientadores limpos
            
    if not cleaned_parts:  # Se nenhuma parte limpa restou
        return "Não Informado"  # Retorna o valor padrão
        
    return "; ".join(cleaned_parts)  # Junta os nomes de orientadores limpos com "; "


def process_record_fields(description: str, advisor: str) -> Tuple[str, str]:
    """
    Limpa e corrige os campos de resumo (description) e orientador (advisor) para evitar dados trocados.
    Se a descrição contiver o nome do orientador, limpa o resumo e usa esse nome para preencher o orientador.
    """
    # 1. Limpa primeiramente a string do orientador enviada
    cleaned_advisor: str = clean_advisor_name(advisor)  # Chama a função de limpeza de orientador
    
    # 2. Verifica se a descrição é na verdade o nome do orientador trocado de campo
    desc_clean: str = description.strip()  # Remove espaços do resumo
    is_misfiled_advisor: bool = False  # Flag indicadora de dado trocado
    extracted_advisor: Optional[str] = None  # Variável para armazenar o orientador extraído da descrição
    
    # Aplica a expressão regular pré-compilada para identificar "Orientador: Nome"
    match = RE_ORIENTADOR_LABEL.match(desc_clean)
    if match:  # Se encontrou o padrão de texto de orientador
        is_misfiled_advisor = True  # Marca que o campo veio trocado
        extracted_advisor = match.group(1).strip()  # Extrai o nome após o rótulo
    elif desc_clean.lower().startswith('orientador'):  # Se apenas começar com a palavra orientador
        if len(desc_clean) < 150:  # E for um texto curto (indicando nome e não um resumo real)
            is_misfiled_advisor = True  # Marca que é dado trocado
            extracted_advisor = desc_clean  # O próprio texto é o orientador
            
    if is_misfiled_advisor:  # Se foi detectado dado trocado
        desc_clean = ""  # Limpa o campo de resumo/descrição
        # Se o orientador não tiver sido informado anteriormente, utiliza o orientador extraído
        if extracted_advisor and (cleaned_advisor == "Não Informado" or len(cleaned_advisor) < len(extracted_advisor)):
            cleaned_advisor = clean_advisor_name(extracted_advisor)  # Limpa e atribui o novo nome
            
    # Normaliza descrições vazias para o texto padrão "Não Informado"
    if not desc_clean:  # Se a descrição ficou vazia
        desc_clean = "Não Informado"  # Define como Não Informado
        
    return desc_clean, cleaned_advisor  # Retorna a tupla com o resumo limpo e o orientador limpo


def create_config_template(file_path: str) -> None:
    """
    Cria um arquivo de template de configuração no formato Excel pré-formatado utilizando openpyxl.
    """
    import openpyxl  # Importa a biblioteca openpyxl para manipular arquivos Excel
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Importa estilos de célula
    
    logger.info(f"Creating a new configuration template: {file_path}")  # Loga a criação do template
    wb = openpyxl.Workbook()  # Cria um novo livro de trabalho Excel
    ws = wb.active  # Pega a planilha ativa principal
    ws.title = "Configuração"  # Renomeia a aba para "Configuração"
    
    # Habilita a exibição das linhas de grade na planilha
    ws.views.sheetView[0].showGridLines = True  # Define linhas de grade visíveis
    
    # Estilos de cores, fontes e bordas para o cabeçalho e conteúdo
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Preenchimento azul escuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")  # Fonte branca em negrito
    label_font = Font(name="Calibri", size=11, bold=True)  # Fonte em negrito para rótulos
    regular_font = Font(name="Calibri", size=11)  # Fonte normal para valores
    
    thin_border = Border(  # Borda cinza clara fina para as células da tabela
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Cabeçalhos da Seção 1 (Parâmetros)
    ws['A1'] = "Configurações Gerais"  # Título da Coluna A
    ws['B1'] = "Valor"  # Título da Coluna B
    ws['A1'].fill = header_fill  # Aplica cor de fundo no cabeçalho A1
    ws['A1'].font = header_font  # Aplica fonte do cabeçalho A1
    ws['B1'].fill = header_fill  # Aplica cor de fundo no cabeçalho B1
    ws['B1'].font = header_font  # Aplica fonte do cabeçalho B1
    
    # Lista de parâmetros de configuração com seus valores padrão
    params: List[Tuple[str, str]] = [
        ("Banco de Dados (SQLite)", "bdtd_metadata.db"),
        ("Excel de Saída", "resultados_coleta.xlsx"),
        ("Limite por Termo (Deixe vazio para coletar todos)", ""),
        ("Atraso entre requisições (segundos)", "2.0"),
        ("Tipo de Busca (Campo de Busca: AllFields, Title, Author, Subject, Advisor)", "AllFields"),
        ("Ordenação (year, year asc, relevance, title, author)", "year"),
        ("Filtro: Tipo de Documento (doctoralThesis, masterThesis, article)", ""),
        ("Filtro: Instituição (Sigla, ex: USP, UNICAMP)", ""),
        ("Filtro: Ano de Publicação (ex: 2025 ou [2020 TO 2026])", ""),
        ("Filtro: Idioma (ex: por, eng, spa)", "")
    ]
    
    # Escreve a lista de parâmetros nas linhas da planilha a partir da linha 2
    for i, (param, val) in enumerate(params, start=2):  # Itera gerando índices de linha
        ws[f'A{i}'] = param  # Define o nome do parâmetro na coluna A
        ws[f'B{i}'] = val  # Define o valor padrão na coluna B
        ws[f'A{i}'].font = label_font  # Aplica a fonte de rótulo
        ws[f'B{i}'].font = regular_font  # Aplica a fonte regular
        ws[f'A{i}'].border = thin_border  # Aplica a borda na coluna A
        ws[f'B{i}'].border = thin_border  # Aplica a borda na coluna B
        
    # Cabeçalhos da Seção 2 (Termos de Busca / Palavras-chave)
    ws['D1'] = "Termos de Busca"  # Título da Coluna D
    ws['D1'].fill = header_fill  # Aplica cor de fundo no cabeçalho D1
    ws['D1'].font = header_font  # Aplica fonte do cabeçalho D1
    
    # Exemplos de termos de busca padrão
    keywords: List[str] = [
        '("Inferencia causal" OR "descoberta causal")',
        'desenvolvimento regional',
        'planejamento urbano'
    ]
    
    # Escreve as palavras-chave na coluna D a partir da linha 2
    for i, kw in enumerate(keywords, start=2):  # Itera escrevendo os termos
        ws[f'D{i}'] = kw  # Define o termo na coluna D
        ws[f'D{i}'].font = regular_font  # Aplica fonte regular
        ws[f'D{i}'].border = thin_border  # Aplica borda na célula
        
    # Ajusta a largura das colunas para melhor visualização
    ws.column_dimensions['A'].width = 55  # Largura da coluna A
    ws.column_dimensions['B'].width = 30  # Largura da coluna B
    ws.column_dimensions['C'].width = 5   # Largura da coluna C (espaçador)
    ws.column_dimensions['D'].width = 50  # Largura da coluna D
    
    wb.save(file_path)  # Salva o arquivo de configuração Excel no caminho indicado
    logger.info(f"Successfully generated config template at {file_path}")  # Loga o sucesso da operação


def read_config_file(file_path: str) -> Dict[str, Any]:
    """
    Lê as configurações de busca e palavras-chave a partir do arquivo Excel de configuração.
    Refatorado para iterar até ws.max_row de forma robusta, ignorando linhas em branco intermediárias.
    """
    import openpyxl  # Importa a biblioteca openpyxl para leitura do Excel
    logger.info(f"Reading configuration from Excel: {file_path}")  # Registra a leitura do arquivo Excel
    
    wb = openpyxl.load_workbook(file_path)  # Carrega a pasta de trabalho Excel
    if "Configuração" not in wb.sheetnames:  # Valida se a aba "Configuração" existe no arquivo
        raise ValueError(f"Sheet 'Configuração' not found in {file_path}")  # Lança erro se a aba não for encontrada
        
    ws = wb["Configuração"]  # Acessa a aba "Configuração"
    
    # Dicionário base com as opções padrão de configuração
    config: Dict[str, Any] = {
        'db_path': 'bdtd_metadata.db',
        'export_excel': 'resultados_coleta.xlsx',
        'limit': None,
        'delay': 2.0,
        'search_type': 'AllFields',
        'sort_order': 'year',
        'filters': [],
        'keywords': []
    }
    
    # Lê os parâmetros das Colunas A (nome) e B (valor) da linha 2 até a linha 19
    for r in range(2, 20):  # Percorre as linhas de parâmetros
        param = ws[f'A{r}'].value  # Lê o rótulo do parâmetro
        val = ws[f'B{r}'].value  # Lê o valor configurado
        if not param:  # Se o rótulo estiver vazio
            continue  # Pula a linha
            
        param_str: str = str(param).strip().lower()  # Converte o rótulo para minúsculas
        if "banco" in param_str:  # Parâmetro de caminho do banco de dados
            if val: config['db_path'] = str(val).strip()  # Atualiza o arquivo .db
        elif "excel" in param_str:  # Parâmetro do arquivo de exportação
            if val: config['export_excel'] = str(val).strip()  # Atualiza o arquivo de saída
        elif "limite" in param_str:  # Parâmetro de limite de registros por termo
            if val is not None and str(val).strip():  # Se houver valor numérico
                try:
                    config['limit'] = int(val)  # Converte para inteiro
                except ValueError:  # Se falhar a conversão
                    logger.warning(f"Invalid limit value: {val}. Defaulting to no limit.")  # Alerta e mantém sem limite
        elif "atraso" in param_str:  # Parâmetro de tempo de espera entre requisições
            if val is not None and str(val).strip():  # Se houver valor
                try:
                    config['delay'] = float(val)  # Converte para float (segundos)
                except ValueError:  # Em caso de valor inválido
                    logger.warning(f"Invalid delay value: {val}. Defaulting to 2.0s.")  # Alerta e mantém 2.0s
        elif "tipo de busca" in param_str:  # Campo de busca Solr (AllFields, Title, etc.)
            if val: config['search_type'] = str(val).strip()  # Define o tipo de busca
        elif "ordenação" in param_str or "ordenacao" in param_str:  # Ordenação dos resultados
            if val: config['sort_order'] = str(val).strip()  # Define a ordenação
        elif "filtro: tipo de documento" in param_str:  # Filtro por formato de documento
            if val:
                fmt: str = str(val).strip()  # Pega a sigla/tipo do formato
                config['filters'].append(f"format:{fmt}")  # Adiciona ao filtro Solr
        elif "filtro: instituição" in param_str or "filtro: instituicao" in param_str:  # Filtro por instituição
            if val:
                inst: str = str(val).strip()  # Sigla da universidade
                config['filters'].append(f"institution:{inst}")  # Adiciona filtro Solr de instituição
        elif "filtro: ano de publicação" in param_str or "filtro: ano de publicacao" in param_str:  # Filtro por ano
            if val:
                yr: str = str(val).strip()  # Ano ou intervalo Solr
                config['filters'].append(f"publishDate:{yr}")  # Adiciona filtro Solr de data
        elif "filtro: idioma" in param_str:  # Filtro por idioma
            if val:
                lang: str = str(val).strip()  # Idioma (ex: por, eng)
                if ',' in lang:  # Se houver múltiplos idiomas separados por vírgula
                    langs: List[str] = [l.strip() for l in lang.split(',')]  # Separa os idiomas
                    for l in langs:  # Para cada idioma
                        config['filters'].append(f"~language:{l}")  # Adiciona com til para pós-processamento local
                else:
                    config['filters'].append(f"language:{lang}")  # Adiciona filtro direto de idioma
                    
    # Lê as palavras-chave da Coluna D a partir da linha 2 até max_row (evita quebrar em células vazias intermediárias)
    for r in range(2, ws.max_row + 1):
        val = ws[f'D{r}'].value  # Lê a célula na coluna D
        if val is not None:  # Se a célula tiver conteúdo
            val_str: str = str(val).strip()  # Remove espaços extras
            if val_str:  # Se a string não for vazia
                config['keywords'].append(val_str)  # Adiciona a palavra-chave à lista
        
    return config  # Retorna o dicionário de configurações completo


def create_json_config_template(file_path: str) -> None:
    """
    Cria um arquivo de modelo de configuração no formato JSON.
    """
    import json  # Importa o módulo nativo para manipular arquivos JSON
    logger.info(f"Creating a new JSON configuration template: {file_path}")  # Loga criação do modelo JSON
    
    # Dicionário com a estrutura e valores padrão para o JSON
    template: Dict[str, Any] = {
        "db_path": "bdtd_metadata.db",
        "export_path": "resultados_coleta.xlsx",
        "limit": None,
        "delay": 2.0,
        "search_type": "AllFields",
        "sort_order": "year",
        "filters": {
            "format": "",
            "institution": "",
            "publishDate": "",
            "language": ""
        },
        "keywords": [
            '("Inferencia causal" OR "descoberta causal")',
            "desenvolvimento regional",
            "planejamento urbano"
        ]
    }
    
    # Grava a estrutura JSON no arquivo com indentação de 4 espaços e suporte a UTF-8
    with open(file_path, 'w', encoding='utf-8') as f:  # Abre o arquivo em modo de escrita
        json.dump(template, f, ensure_ascii=False, indent=4)  # Escreve o JSON formatado
    logger.info(f"Successfully generated JSON config template at {file_path}")  # Loga conclusão


def read_json_config_file(file_path: str) -> Dict[str, Any]:
    """
    Lê as configurações de busca e palavras-chave do arquivo JSON com tratamento específico de erros Pydantic.
    """
    logger.info(f"Reading configuration from JSON: {file_path}")  # Loga a leitura do arquivo JSON
    
    try:
        from config_app.core.config_schemas import BDTDConfig, load_and_validate_config  # Importa o validador
        validated = load_and_validate_config(file_path, BDTDConfig)  # Valida o arquivo contra o modelo BDTDConfig
        raw_data: Dict[str, Any] = validated.model_dump()  # Converte o objeto validado de volta para um dicionário
    except (ImportError, ModuleNotFoundError):  # Tratamento específico para ausência da biblioteca Pydantic / Schema
        logger.warning("Schemas Pydantic não encontrados. Usando fallback de leitura bruta de JSON.")
        import json  # Importa o módulo nativo de JSON
        with open(file_path, 'r', encoding='utf-8') as f:  # Abre o arquivo JSON em modo leitura
            raw_data = json.load(f)  # Carrega o dicionário bruto

    # Constrói a estrutura de configuração padronizada
    config: Dict[str, Any] = {
        'db_path': raw_data.get('db_path', 'bdtd_metadata.db'),  # Caminho do banco SQLite
        'export_excel': raw_data.get('export_path', 'resultados_coleta.xlsx'),  # Caminho do arquivo exportado
        'limit': raw_data.get('limit'),  # Limite de registros por termo
        'delay': float(raw_data.get('delay', 2.0)),  # Atraso entre chamadas em segundos
        'search_type': raw_data.get('search_type', 'AllFields'),  # Tipo de campo de busca
        'sort_order': raw_data.get('sort_order', 'year'),  # Ordenação dos resultados Solr
        'filters': [],  # Lista de filtros formatados
        'scrape_details': raw_data.get('scrape_details', True),  # Flag para ativar raspagem de página web
        'keywords': raw_data.get('keywords', [])  # Lista de palavras-chave
    }
    
    # Processa os filtros definidos no arquivo JSON
    filters_dict = raw_data.get('filters', {})  # Obtém o bloco de filtros
    if isinstance(filters_dict, dict):  # Se for um dicionário de pares chave:valor
        for key, val in filters_dict.items():  # Percorre cada filtro configurado
            if val:  # Se o valor do filtro não for vazio
                val_str: str = str(val).strip()  # Limpa o texto
                if key == 'language' and ',' in val_str:  # Tratamento para múltiplos idiomas
                    langs: List[str] = [l.strip() for l in val_str.split(',')]  # Separa os idiomas por vírgula
                    for lang in langs:  # Para cada idioma
                        config['filters'].append(f"~language:{lang}")  # Adiciona formato de til para pós-filtro local
                else:
                    config['filters'].append(f"{key}:{val_str}")  # Adiciona filtro padrão campo:valor
    elif isinstance(filters_dict, list):  # Se já for uma lista pronta de filtros
        config['filters'] = filters_dict  # Usa a lista diretamente
            
    return config  # Retorna a configuração final tratada


def extract_authors(authors_dict: Optional[Dict[str, Any]]) -> str:
    """
    Extrai e consolida todos os autores primários, secundários e corporativos em uma única string.
    Trata diferentes estruturas de dados possíveis no dicionário de autores retornado pela API.
    """
    if not authors_dict:  # Se o dicionário de autores for nulo ou vazio
        return ""  # Retorna string vazia
    
    author_list: List[str] = []  # Lista para armazenar o nome de todos os autores extraídos
    
    # 1. Autores Primários (Autor principal do trabalho)
    primary = authors_dict.get('primary', {})  # Obtém bloco de autores primários
    if isinstance(primary, dict):  # Se vier como dicionário onde as chaves são os nomes
        author_list.extend(primary.keys())  # Adiciona os nomes (chaves) à lista
    elif isinstance(primary, list):  # Se vier como uma lista de nomes
        author_list.extend(primary)  # Adiciona a lista inteira
        
    # 2. Autores Secundários (Coautores, orientadores, coorientadores)
    secondary = authors_dict.get('secondary', [])  # Obtém bloco de autores secundários
    if isinstance(secondary, dict):  # Se vier como dicionário
        author_list.extend(secondary.keys())  # Adiciona as chaves
    elif isinstance(secondary, list):  # Se vier como lista
        for sa in secondary:  # Itera sobre cada elemento da lista
            if isinstance(sa, dict):  # Se o elemento for um dicionário
                author_list.extend(sa.keys())  # Adiciona as chaves
            else:
                author_list.append(str(sa))  # Adiciona a representação em texto do item
                
    # 3. Autores Corporativos (Instituições, empresas, comissões)
    corporate = authors_dict.get('corporate', [])  # Obtém bloco de autores corporativos
    if isinstance(corporate, dict):  # Se for dicionário
        author_list.extend(corporate.keys())  # Adiciona as chaves
    elif isinstance(corporate, list):  # Se for lista
        for ca in corporate:  # Percorre cada elemento
            if isinstance(ca, dict):  # Se o item for um dicionário
                author_list.extend(ca.keys())  # Adiciona as chaves
            else:
                author_list.append(str(ca))  # Converte e adiciona como texto
                
    return "; ".join([a.strip() for a in author_list if a])  # Retorna todos os autores únicos unidos por "; "


def extract_subjects(subjects_list: Optional[List[Any]]) -> str:
    """
    Extrai os assuntos (palavras-chave) da lista de listas retornada pela API Solr e une por ponto e vírgula.
    Exemplo: [['Ciências Humanas - Geografia'], ['Sustentabilidade']] -> "Ciências Humanas - Geografia; Sustentabilidade"
    """
    if not subjects_list:  # Se a lista de assuntos for nula ou vazia
        return ""  # Retorna string vazia
    
    keywords: List[str] = []  # Lista para armazenar as palavras-chave tratadas
    for sub in subjects_list:  # Percorre cada item retornado
        if isinstance(sub, list) and len(sub) > 0:  # Se o item for uma lista aninhada com conteúdo
            keywords.append(str(sub[0]).strip())  # Pega o primeiro elemento da sublista
        elif isinstance(sub, str):  # Se for uma string direta
            keywords.append(sub.strip())  # Limpa e adiciona à lista
            
    return "; ".join([k for k in keywords if k])  # Junta os assuntos válidos utilizando "; "


def extract_url(urls_list: Optional[List[Any]]) -> str:
    """
    Extrai a string de URL da lista de URLs retornada pela API.
    Trata tanto dicionários (ex: [{'url': '...', 'desc': '...'}]) quanto strings diretas.
    """
    if not urls_list:  # Se a lista de URLs for vazia ou nula
        return ""  # Retorna string vazia
    first_url = urls_list[0]  # Pega a primeira URL da lista
    if isinstance(first_url, dict):  # Se o item for um dicionário com chave 'url'
        return first_url.get('url', '').strip()  # Extrai a URL limpa do dicionário
    return str(first_url).strip()  # Caso contrário, converte o item para string e limpa


def translate_format(fmt_str: Optional[str]) -> str:
    """Traduz a chave de tipo de formato do VuFind para uma descrição amigável em português."""
    if not fmt_str:  # Se a string de formato for nula ou vazia
        return "Tese/Dissertação"  # Retorna valor amigável padrão
    fmt: str = fmt_str.lower()  # Converte o formato para minúsculas para comparação
    if 'doctoralthesis' in fmt or 'doctoral thesis' in fmt or 'tese' in fmt:
        return 'Tese'  # Identifica e traduz como Tese de Doutorado
    elif 'masterthesis' in fmt or 'master thesis' in fmt or 'dissertacao' in fmt or 'dissertação' in fmt:
        return 'Dissertação'  # Identifica e traduz como Dissertação de Mestrado
    elif 'article' in fmt or 'artigo' in fmt or 'journal' in fmt:
        return 'Artigo'  # Identifica e traduz como Artigo Científico
    elif 'book' in fmt or 'livro' in fmt:
        return 'Livro'  # Identifica e traduz como Livro
    elif 'chapter' in fmt or 'capitulo' in fmt:
        return 'Capítulo de Livro'  # Identifica e traduz como Capítulo de Livro
    return fmt_str.capitalize()  # Caso seja outro formato, apenas capitaliza a primeira letra


def scrape_record_details(record_id: str, session: Optional[requests.Session] = None, cookie: str = "OasisbrVerify=verified_human") -> Dict[str, str]:
    """
    Realiza a raspagem (scraping) da página pública da BDTD para um único registro
    utilizando requests.Session com conexão persistente (HTTP Keep-Alive) para máxima velocidade.
    """
    url: str = f"https://bdtd.ibict.br/vufind/Record/{record_id}"  # Monta a URL pública da página do registro
    headers: Dict[str, str] = {
        'Cookie': cookie,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'
    }
    
    details: Dict[str, str] = {}  # Dicionário que armazenará todos os pares de metadados raspados
    try:
        # Reutiliza a sessão HTTP ativa para manter a conexão TCP aberta (Keep-Alive)
        if session:
            res = session.get(url, headers=headers, timeout=30)
        else:
            res = requests.get(url, headers=headers, timeout=30)
        res.raise_for_status()  # Lança exceção em caso de status de erro HTTP (ex: 404, 500)
        
        html: str = res.text  # Obtém o conteúdo HTML em texto
        soup: BeautifulSoup = BeautifulSoup(html, 'html.parser')  # Instancia o analisador HTML do BeautifulSoup
        
        # 1. Raspa tags de metadados Dublin Core (<meta name="..." content="...">) no cabeçalho <head>
        for meta in soup.find_all('meta'):  # Procura todas as tags <meta> do HTML
            name = meta.attrs.get('name')  # Obtém o atributo name
            content = meta.attrs.get('content')  # Obtém o atributo content
            if name and content:  # Se ambos existirem
                details[name] = content  # Guarda a informação no dicionário de detalhes
                
        # 2. Raspa as tabelas HTML de metadados exibidas na aba de detalhes do registro
        for row in soup.find_all('tr'):  # Procura todas as linhas <tr> de tabelas na página
            th = row.find('th')  # Localiza o cabeçalho da linha (rótulo do campo)
            td = row.find('td')  # Localiza a célula de dados (valor do campo)
            if th and td:  # Se encontrou ambos os elementos na linha
                key: str = th.text.strip()  # Extrai o nome do campo limpo
                val: str = td.text.strip()  # Extrai o valor do campo limpo
                # Caso o valor possua múltiplas linhas (listas de assuntos, autores ou editoras)
                lines: List[str] = [l.strip() for l in val.split('\n') if l.strip()]  # Divide em linhas não vazias
                if len(lines) > 1:  # Se houver mais de uma linha
                    details[key] = "; ".join(lines)  # Junta tudo com ponto e vírgula
                else:
                    details[key] = lines[0] if lines else ""  # Atribui o valor único ou string vazia
    except Exception as e:  # Caso ocorra falha na requisição ou parsing da página
        logger.warning(f"Failed to scrape webpage details for record {record_id}: {e}")  # Emite alerta no log
        
    return details  # Retorna o dicionário com os detalhes raspados da página web


def get_source_info(details: Dict[str, str]) -> str:
    """
    Extrai o contexto da fonte (Universidade, Editora ou Revista)
    com base no formato do documento e nos campos disponíveis nos detalhes.
    """
    fmt: str = details.get('format', '').lower()  # Obtém a string de formato em minúsculas
    
    # Lista de campos candidatos para identificação de instituição, editora e fonte
    institution = details.get('instname_str') or details.get('Instituição de defesa:') or details.get('Instituição:') or details.get('institution')
    publisher = details.get('publisher.none.fl_str_mv') or details.get('dc.publisher.none.fl_str_mv') or details.get('Editora:') or details.get('Editora')
    source = details.get('reponame_str') or details.get('dc.source.none.fl_str_mv') or details.get('Fonte:') or details.get('container_title')
    
    # Para Teses e Dissertações, prioriza a Instituição de Defesa
    if 'thesis' in fmt or 'tese' in fmt or 'dissertacao' in fmt:
        return institution or source or publisher or "Não Informado"
    # Para Livros, prioriza a Editora
    elif 'book' in fmt or 'livro' in fmt:
        return publisher or institution or source or "Não Informado"
    # Para Artigos Científicos, prioriza a Revista / Fonte
    elif 'article' in fmt or 'artigo' in fmt or 'journal' in fmt:
        return source or publisher or institution or "Não Informado"
        
    # Fallback genérico caso não se enquadre nas categorias anteriores
    return institution or publisher or source or "Não Informado"


class DatabaseManager:
    """
    Gerencia conexões com o banco de dados SQLite, criação da tabela de esquema,
    inserções individuais e em lote com suporte a UPSERT e transações eficientes.
    """
    def __init__(self, db_path: str) -> None:
        self.db_path: str = db_path  # Salva o caminho do arquivo de banco de dados SQLite
        self.conn: Optional[sqlite3.Connection] = None  # Inicializa o atributo de conexão como None
        self.init_db()  # Executa a inicialização do banco e tabelas

    def init_db(self) -> None:
        """Inicializa o esquema do banco de dados e cria as tabelas e índices se não existirem."""
        try:
            open_db_path: str = self.db_path  # Obtém o caminho do banco de dados
            # Trata limitação de tamanho de caminho no Windows (Long Path Support)
            if sys.platform.startswith('win') and len(os.path.abspath(open_db_path)) >= 240 and not os.path.abspath(open_db_path).startswith('\\\\?\\'):
                open_db_path = '\\\\?\\' + os.path.abspath(open_db_path)  # Adiciona prefixo do Windows para caminhos longos
            self.conn = sqlite3.connect(open_db_path)  # Abre a conexão com o arquivo do banco SQLite
            cursor = self.conn.cursor()  # Obtém o cursor para executar comandos SQL
            
            # Cria a tabela de metadados acadêmicos com tipagens SQL corretas se ela não existir
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS academic_metadata (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    record_id TEXT UNIQUE,
                    title TEXT,
                    creator TEXT,
                    date TEXT,
                    description TEXT,
                    subject TEXT,
                    type_of_research TEXT,
                    advisor TEXT,
                    source_institution TEXT,
                    download_url TEXT,
                    harvested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            
            # Cria índice no campo record_id para verificações ultra-rápidas de duplicatas
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_record_id ON academic_metadata (record_id);")
            # Cria índice na data de publicação para otimização de consultas por ano
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_date ON academic_metadata (date);")
            
            self.conn.commit()  # Confirma a transação SQL no banco
            logger.info(f"Database initialized successfully at: {self.db_path}")  # Loga sucesso da inicialização
        except sqlite3.Error as e:  # Captura exceções específicas de erro SQLite
            logger.error(f"Failed to initialize SQLite database: {e}")  # Registra o erro no log
            raise  # Lança novamente a exceção para interromper o pipeline em caso de erro grave

    def insert_batch(self, records_list: List[Dict[str, Any]]) -> int:
        """
        Insere ou atualiza (UPSERT) uma lista de registros no banco de dados SQLite.
        Utiliza transação única (executemany) para evitar fsync a cada registro, aumentando a velocidade em até 100x.
        Atualiza dados se o registro já existir e a nova versão trouxer informações mais completas.
        """
        if not records_list or not self.conn:  # Se a lista enviada estiver vazia ou conexão indisponível
            return 0  # Retorna 0 inserções
            
        # Cláusula UPSERT nativa do SQLite para atualização se já existir o record_id
        query = """
            INSERT INTO academic_metadata 
            (record_id, title, creator, date, description, subject, type_of_research, advisor, source_institution, download_url)
            VALUES (:record_id, :title, :creator, :date, :description, :subject, :type_of_research, :advisor, :source_institution, :download_url)
            ON CONFLICT(record_id) DO UPDATE SET
                title = excluded.title,
                creator = excluded.creator,
                date = excluded.date,
                description = CASE WHEN excluded.description != 'Não Informado' THEN excluded.description ELSE academic_metadata.description END,
                subject = excluded.subject,
                type_of_research = excluded.type_of_research,
                advisor = CASE WHEN excluded.advisor != 'Não Informado' THEN excluded.advisor ELSE academic_metadata.advisor END,
                source_institution = CASE WHEN excluded.source_institution != 'Não Informado' THEN excluded.source_institution ELSE academic_metadata.source_institution END,
                download_url = excluded.download_url;
        """
        try:
            cursor = self.conn.cursor()  # Obtém o cursor
            cursor.executemany(query, records_list)  # Executa o lote inteiro na transação
            self.conn.commit()  # Um único commit para todo o lote de registros
            return cursor.rowcount  # Retorna o número de registros processados
        except sqlite3.Error as e:  # Caso ocorra erro durante a inserção do lote
            logger.error(f"Error inserting batch into SQLite database: {e}")  # Loga a falha
            if self.conn:
                self.conn.rollback()  # Desfaz a transação em caso de erro
            return 0

    def insert_record(self, record_dict: Dict[str, Any]) -> bool:
        """
        Insere ou atualiza um único registro acadêmico no banco de dados.
        Refatorado para utilizar a lógica em lote de alta performance (insert_batch).
        """
        return self.insert_batch([record_dict]) > 0  # Retorna True se a operação afetou a tabela

    def close(self) -> None:
        """Encerra com segurança a conexão ativa com o banco de dados SQLite."""
        if self.conn:  # Se houver conexão ativa
            self.conn.close()  # Fecha o banco de dados
            self.conn = None
            logger.info("Database connection closed.")  # Registra encerramento no log


class BDTDHarvesterPipeline:
    """
    Pipeline principal que executa buscas na BDTD, processa registros encontrados,
    realiza raspagem de detalhes via requests.Session (Keep-Alive) e coordena a persistência otimizada no SQLite.
    """
    def __init__(
        self, 
        db_manager: DatabaseManager, 
        keywords: List[str], 
        limit_per_keyword: Optional[int] = None, 
        delay: float = 2.0, 
        page_size: int = 100,
        search_type: str = 'AllFields', 
        sort_order: str = 'year', 
        filters: Optional[List[str]] = None, 
        scrape_details: bool = True
    ) -> None:
        self.db: DatabaseManager = db_manager  # Instância do gerenciador do banco de dados
        self.keywords: List[str] = keywords  # Lista de palavras-chave para pesquisar
        self.limit_per_keyword: Optional[int] = limit_per_keyword  # Limite máximo de registros por palavra-chave
        self.delay: float = delay  # Tempo de atraso entre chamadas da API
        self.page_size: int = page_size  # Quantidade de resultados por página de busca
        self.search_type: str = search_type or 'AllFields'  # Tipo do campo de busca Solr
        self.sort_order: str = sort_order or 'year'  # Ordenação dos resultados
        self.scrape_details: bool = scrape_details  # Flag booleana indicando se deve raspar página web
        
        # Inicializa a sessão HTTP do requests com cabeçalhos padrão (Keep-Alive)
        self.session: requests.Session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
            'Cookie': 'OasisbrVerify=verified_human'
        })
        
        # Sanitiza os filtros aplicando as regras de contorno do WAF da BDTD
        self.api_filters: List[str]
        self.allowed_languages: Set[str]
        self.api_filters, self.allowed_languages = sanitize_bdtd_filters(filters)
        
        if self.allowed_languages:  # Se houver filtros de idioma definidos
            logger.info(f"Language filter will be applied locally (post-filter): {self.allowed_languages}")
        if self.api_filters:  # Se houver filtros da API configurados
            logger.info(f"API filters to send ({len(self.api_filters)}/{BDTD_MAX_API_FILTERS} max): {self.api_filters}")
        
        # Lista de campos explicitamente solicitados à API do VuFind
        self.request_fields: List[str] = [
            'id', 'title', 'authors', 'subjects', 
            'languages', 'formats', 'urls', 'summary', 'publicationDates'
        ]

    def fetch_with_retries(self, keyword: str, page: int, max_retries: int = 5, backoff_factor: float = 2.0) -> Dict[str, Any]:
        """
        Consulta bdtd_scraper.api.get_search_results com lógica de tentativas (retries) e backoff exponencial para estabilidade.
        """
        retries: int = 0  # Contador de tentativas realizadas
        current_delay: float = self.delay  # Atraso base configurado
        
        # Limpa o termo de busca para evitar bloqueios no WAF (erros HTTP 429) e buscas por proximidade muito lentas no Lucene.
        clean_keyword: str = keyword.replace('"', '').replace("'", "").strip()  # Remove aspas duplas e simples
        clean_keyword = unicodedata.normalize('NFKD', clean_keyword).encode('ASCII', 'ignore').decode('utf-8')  # Remove acentuação
        if clean_keyword != keyword:  # Se a palavra limpa for diferente da original
            logger.info(f"Cleaned query keyword: from '{keyword}' to '{clean_keyword}'")  # Registra no log
            
        # Monta os parâmetros de consulta para a API
        query_params: Dict[str, Any] = {
            'lookfor': clean_keyword,
            'page': page,
            'limit': self.page_size,
            'field': self.request_fields,
            'type': self.search_type,
            'sort': self.sort_order
        }
        
        if self.api_filters:  # Se houver filtros de API válidos
            query_params['filter'] = self.api_filters  # Adiciona os filtros na consulta
            
        while retries < max_retries:  # Loop de tentativas
            is_rate_limited: bool = False  # Flag para identificar bloqueio por excesso de requisições (429)
            try:
                logger.info(f"Querying BDTD for '{keyword}' - Page {page}...")  # Registra chamada no log
                response: Dict[str, Any] = bdtd_scraper.api.get_search_results(**query_params)  # Chama o método da biblioteca bdtd_scraper
                
                if response.get('status') == 'OK' or 'records' in response:  # Se a resposta foi bem-sucedida
                    # Aplica a filtragem local por idioma, se configurada
                    if self.allowed_languages and 'records' in response:  # Se houver filtro de idioma ativo
                        original_count: int = len(response['records'])  # Contagem original de itens retornado
                        response['records'] = [  # Filtra localmente mantendo registros que correspondem ao idioma
                            rec for rec in response['records']
                            if not rec.get('languages') or
                               any(lang in self.allowed_languages for lang in rec.get('languages', []))
                        ]
                        filtered_count: int = original_count - len(response['records'])  # Quantidade descartada pelo filtro local
                        if filtered_count > 0:  # Se algum registro foi descartado
                            logger.info(f"Local language filter removed {filtered_count} records (kept {len(response['records'])}/{original_count})")
                    return response  # Retorna a resposta válida da API
                
                logger.warning(f"Unexpected status in API response: {response.get('status')}")  # Registra status não esperado
                
            except Exception as e:  # Captura falhas de conexão ou HTTP
                err_msg: str = str(e)  # Converte o erro para texto
                logger.warning(f"Error querying API (Attempt {retries + 1}/{max_retries}): {err_msg}")  # Loga tentativa incorreta
                if "429" in err_msg:  # Se for erro 429 de limite de requisição
                    is_rate_limited = True  # Ativa o tratamento específico de rate-limiting
                
            retries += 1  # Incrementa o número de tentativas
            if retries < max_retries:  # Se ainda restarem tentativas
                sleep_time: float = current_delay * (backoff_factor ** (retries - 1))  # Calcula o tempo de espera com backoff exponencial
                if is_rate_limited:  # Se foi bloqueado pelo WAF por limite
                    sleep_time = max(15.0, sleep_time)  # Garante pausa mínima de 15 segundos para desbloqueio
                    logger.warning(f"Rate limit hit (429). Cool down sleep initiated for {sleep_time:.2f} seconds...")
                else:
                    logger.info(f"Retrying in {sleep_time:.2f} seconds...")  # Loga o tempo de espera simples
                time.sleep(sleep_time)  # Executa a pausa na thread
                
        raise ConnectionError(f"Failed to fetch data from BDTD API after {max_retries} attempts.")  # Interrompe se esgotar as tentativas

    def run(self) -> None:
        """Executa o pipeline completo de coleta de dados utilizando lote (batch_buffer) e requests.Session."""
        total_inserted: int = 0  # Contador total de registros inseridos no banco
        total_processed: int = 0  # Contador total de registros processados
        
        # Conjunto para controle local de unicidade de registros nesta execução
        seen_record_ids: Set[str] = set()

        logger.info(f"Starting metadata harvest for keywords: {self.keywords}")  # Loga o início do processo
        start_time: float = time.time()  # Registra o tempo inicial de execução

        for keyword in self.keywords:  # Itera sobre cada palavra-chave da lista
            logger.info(f"Processing keyword: '{keyword}'")  # Loga a palavra-chave atual
            page: int = 1  # Reinicia o número da página para 1
            keyword_inserted: int = 0  # Contador de inserções para a palavra-chave atual
            keyword_processed: int = 0  # Contador de processamentos para a palavra-chave atual
            has_more_results: bool = True  # Flag de controle de paginação
            
            while has_more_results:  # Enquanto houver páginas a buscar
                try:
                    res: Dict[str, Any] = self.fetch_with_retries(keyword, page)  # Busca a página de resultados da API
                except Exception as e:  # Se falhar todas as tentativas da página
                    logger.error(f"Skipping keyword '{keyword}' due to API connection failure: {e}")
                    break  # Abandona a palavra-chave atual e passa para a próxima
                
                records: List[Dict[str, Any]] = res.get('records', [])  # Extrai a lista de registros retornados
                result_count: int = res.get('resultCount', 0)  # Extrai o total geral de resultados na BDTD
                
                if not records:  # Se a página retornou sem registros
                    logger.info(f"No records returned for '{keyword}' on page {page}. Stopping.")
                    break  # Encerra a busca para esta palavra-chave
                
                logger.info(f"Retrieved {len(records)} records (Total matching on BDTD: {result_count})")
                
                batch_buffer: List[Dict[str, Any]] = []  # Buffer em memória para acumular registros e salvar em lote por página
                
                for record in records:  # Percorre cada registro retornado da página
                    record_id: Optional[str] = record.get('id')  # Obtém o ID único do registro no VuFind
                    if not record_id:  # Se não houver ID válido
                        continue  # Ignora o registro
                        
                    keyword_processed += 1  # Incrementa registros processados da palavra-chave
                    total_processed += 1  # Incrementa registros processados globais
                    
                    # Deduplicação no conjunto local para evitar reprocessamento no mesmo loop
                    if record_id in seen_record_ids:
                        continue  # Pula se o ID já tiver sido tratado nesta execução
                    seen_record_ids.add(record_id)  # Registra o ID como já visto
                    
                    # Extrai os campos principais com fallbacks seguros e funções auxiliares de tratamento
                    title: str = record.get('title', '').strip()  # Extrai e limpa o título
                    creator: str = extract_authors(record.get('authors'))  # Extrai os autores do registro
                    
                    pub_dates: List[Any] = record.get('publicationDates', [])  # Obtém a lista de datas de publicação
                    date: str = str(pub_dates[0]).strip() if pub_dates else ""  # Pega o primeiro ano retornado
                    
                    summaries: List[str] = record.get('summary', [])  # Obtém a lista de resumos
                    description: str = " ".join([s.strip() for s in summaries if s]).strip()  # Junta todos os resumos
                    
                    subject: str = extract_subjects(record.get('subjects'))  # Extrai as palavras-chave/assuntos
                    
                    # Realiza a raspagem (scraping) da página web pública utilizando a sessão do requests (Keep-Alive)
                    details: Dict[str, str]
                    if getattr(self, 'scrape_details', True):  # Se o scrape de detalhes estiver ativo
                        logger.info(f" -> Scraping page details for record: {record_id}...")
                        details = scrape_record_details(record_id, session=self.session)  # Raspa usando a sessão HTTP
                        # Pausa de cortesia de 1 segundo entre requisições de raspagem para não sobrecarregar o servidor
                        time.sleep(1.0)
                    else:
                        details = {}  # Mantém dicionário de detalhes vazio caso esteja no modo rápido
                    
                    # Tradução e unificação do tipo de pesquisa / formato do trabalho
                    type_of_research: str = translate_format(
                        details.get('format') or 
                        (record.get('formats', [''])[0] if record.get('formats') else '')
                    )
                    
                    # Lógica para extração e validação do nome do orientador
                    advisor: Optional[str] = details.get('Orientador(a):')  # Busca orientação na página web
                    if not advisor or "não informado" in advisor.lower():  # Se estiver nulo ou genérico
                        advisor = details.get('dc.contributor.none.fl_str_mv') or advisor  # Tenta pegar dos metadados Solr
                    if not advisor:  # Se continuou vazio
                        advisor = "Não Informado"  # Define o valor padrão
                        
                    # Extrai a instituição de origem / universidade / editora / revista
                    source_institution: str = get_source_info(details)  # Identifica a instituição pela página web
                    if source_institution == "Não Informado" and record.get('institutions'):  # Se não achou na web mas tem na API
                        source_institution = record.get('institutions')[0]  # Usa o valor da API Solr
                        
                    # Limpa nomes de autores descartando anos de nascimento/morte
                    creator = clean_creator_name(creator)
                    
                    # Corrige dados trocados entre resumo (description) e orientador (advisor) - assinatura simplificada
                    description, advisor = process_record_fields(description=description, advisor=advisor)
                    
                    # Extração robusta do link de download/acesso direto
                    download_url: str = extract_url(record.get('urls'))  # Extrai da API
                    if not download_url and details.get('Link de acesso:'):  # Se falhar, tenta da raspagem web
                        download_url = details.get('Link de acesso:')
                    if not download_url and details.get('url'):  # Outra opção de tag de página web
                        download_url = details.get('url')
                    if not download_url:  # Fallback final se nenhum link for encontrado
                        download_url = f"https://bdtd.ibict.br/vufind/Record/{record_id}"  # Link da página no VuFind
                    
                    # Constrói o dicionário final do registro limpo
                    record_dict: Dict[str, Any] = {
                        'record_id': record_id,
                        'title': title,
                        'creator': creator,
                        'date': date,
                        'description': description,
                        'subject': subject,
                        'type_of_research': type_of_research,
                        'advisor': advisor,
                        'source_institution': source_institution,
                        'download_url': download_url
                    }
                    
                    batch_buffer.append(record_dict)  # Adiciona ao lote da página atual

                # Descarrega o lote da página no banco SQLite em uma transação única (UPSERT)
                if batch_buffer:
                    inserted_count: int = self.db.insert_batch(batch_buffer)  # Salva o lote em uma única escrita
                    keyword_inserted += inserted_count  # Incrementa salvos do termo
                    total_inserted += inserted_count  # Incrementa salvos globais
                    logger.info(f" -> [BATCH SAVED] Saved/Updated batch of {len(batch_buffer)} records in SQLite.")
                    batch_buffer.clear()  # Limpa o buffer de memória do lote
                        
                # Verifica se atingiu o limite de salvamentos configurado após o lote da página
                if self.limit_per_keyword and keyword_inserted >= self.limit_per_keyword:
                    logger.info(f"Limit of {self.limit_per_keyword} records reached for keyword '{keyword}'.")
                    break  # Encerra o loop da palavra-chave
                    
                # Condição de encerramento de paginação da busca
                if page * self.page_size >= result_count or len(records) < self.page_size:
                    has_more_results = False  # Marca que não há mais páginas
                    logger.info(f"All BDTD pages exhausted for keyword '{keyword}'.")
                else:
                    page += 1  # Avança para a próxima página
                    time.sleep(self.delay)  # Respeita o atraso configurado entre chamadas da API
                    
            logger.info(f"Finished '{keyword}': processed {keyword_processed} records, saved {keyword_inserted} relevant records.")

        elapsed: float = time.time() - start_time  # Calcula o tempo total decorrido no pipeline
        logger.info(f"Pipeline execution completed in {elapsed:.2f} seconds.")  # Loga a duração total
        logger.info(f"Total processed: {total_processed} | Total saved to DB: {total_inserted}")  # Loga resumo final


def export_to_format(db_path: str, export_path: str, chunksize: int = 50000) -> bool:
    """
    Exporta os dados da tabela academic_metadata do banco SQLite para Excel, CSV ou JSON.
    Otimizado para alto desempenho e baixo uso de memória (evita estouros de RAM / OOM).
    """
    logger.info(f"Exporting database records to: {export_path}")  # Loga o início da exportação
    try:
        conn: sqlite3.Connection = sqlite3.connect(db_path)  # Abre a conexão com o banco SQLite
        
        # Consulta SQL estruturada renomeando as colunas conforme a exigência do usuário
        query: str = """
            SELECT 
                creator AS "Autores",
                title AS "Título",
                date AS "Ano",
                type_of_research AS "Tipo de Pesquisa",
                advisor AS "Nome do Orientador",
                source_institution AS "Universidade / Editora / Revista",
                description AS "Resumo",
                download_url AS "Link para Download"
            FROM academic_metadata
            ORDER BY harvested_at DESC
        """
        
        ext: str = os.path.splitext(export_path)[1].lower()  # Identifica a extensão do arquivo de saída
        
        # Leitura e gravação otimizadas por blocos (chunksize) para evitar estouro de memória (OOM)
        if ext == '.csv':
            first_chunk: bool = True
            total_records: int = 0
            for chunk_df in pd.read_sql_query(query, conn, chunksize=chunksize):
                chunk_df.to_csv(export_path, mode='a' if not first_chunk else 'w', index=False, header=first_chunk, encoding='utf-8')
                first_chunk = False
                total_records += len(chunk_df)
            conn.close()
            if total_records == 0:
                logger.warning("No records found in the database to export.")
                return False
            logger.info(f"Successfully exported {total_records} records to CSV.")
            return True
            
        # Para Excel e JSON, lê os dados através do cursor SQLite otimizado
        df: pd.DataFrame = pd.read_sql_query(query, conn)
        conn.close()
        
        if df.empty:  # Se não houver registros salvos no banco
            logger.warning("No records found in the database to export.")  # Emite alerta
            return False  # Retorna False
            
        if ext in ['.xlsx', '.xls']:  # Se for planilha Excel
            df.to_excel(export_path, index=False)  # Salva em arquivo Excel sem a coluna de índice
        elif ext == '.json':  # Se for formato JSON
            df.to_json(export_path, orient='records', force_ascii=False, indent=4)  # Salva em JSON formatado
        else:  # Caso a extensão não seja reconhecida
            logger.warning(f"Unrecognized export format: {ext}. Defaulting to CSV.")
            df.to_csv(export_path, index=False, encoding='utf-8')  # Utiliza o padrão CSV
            
        logger.info(f"Successfully exported {len(df)} records.")  # Loga a quantidade de linhas exportadas
        return True  # Retorna True indicando exportação bem-sucedida
    except Exception as e:  # Em caso de falha na exportação
        logger.error(f"Failed to export database: {e}")  # Loga a mensagem de erro
        return False  # Retorna False


def run_harvest(
    keywords: List[str], 
    db_path: str = 'bdtd_metadata.db', 
    export_path: str = 'bdtd_metadata.xlsx', 
    limit: Optional[int] = None, 
    delay: float = 2.0, 
    page_size: int = 100, 
    search_type: str = 'AllFields', 
    sort_order: str = 'year', 
    filters: Optional[List[str]] = None, 
    scrape_details: bool = True
) -> bool:
    """
    Ponto de entrada programático de alto nível para executar a coleta de metadados da BDTD.
    """
    log_dir: str = os.path.dirname(export_path) if os.path.dirname(export_path) else "."
    log_file_path: str = os.path.join(log_dir, "log_bdtd.txt")
    setup_logging(log_file_path)  # Configura logging limpo para arquivo e console

    db_manager: Optional[DatabaseManager] = None  # Inicializa a variável do gerenciador de banco como None
    try:
        # 1. Inicializa o gerenciador do banco de dados SQLite
        db_manager = DatabaseManager(db_path)
        
        # 2. Instancia e executa o Pipeline de Coleta BDTD
        pipeline: BDTDHarvesterPipeline = BDTDHarvesterPipeline(
            db_manager=db_manager,
            keywords=keywords,
            limit_per_keyword=limit,
            delay=delay,
            page_size=page_size,
            search_type=search_type,
            sort_order=sort_order,
            filters=filters,
            scrape_details=scrape_details
        )
        pipeline.run()  # Executa a coleta dos registros
        
        # 3. Fecha a conexão para garantir gravação de todos os dados pendentes no disco
        db_manager.close()
        db_manager = None  # Limpa a referência da variável
        
        # 4. Exporta a tabela do banco SQLite para o formato Excel/CSV/JSON solicitado
        export_to_format(db_path, export_path)
        return True  # Retorna True em caso de sucesso total
    except Exception as e:  # Captura falhas durante a execução do pipeline
        logger.error(f"Pipeline execution failed: {e}")  # Registra o erro no log
        return False  # Retorna False indicando que o pipeline falhou
    finally:
        if db_manager:  # Bloco final que garante o fechamento do banco mesmo se ocorrer um erro
            db_manager.close()


def main() -> None:
    """Função principal para execução do script a partir da linha de comando (CLI)."""
    parser = argparse.ArgumentParser(  # Instancia o analisador de argumentos CLI
        description="Automated Python pipeline to extract BDTD academic metadata and save to SQLite / Excel.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    # Definições de todos os argumentos suportados na linha de comando
    parser.add_argument(
        '--db-path', 
        type=str, 
        default='bdtd_metadata.db', 
        help="Filename or path of the destination SQLite database."
    )
    parser.add_argument(
        '--limit', 
        type=int, 
        default=None, 
        help="Maximum number of records to retrieve and save PER keyword. If omitted, collects all matching records."
    )
    parser.add_argument(
        '--delay', 
        type=float, 
        default=2.0, 
        help="Polite wait time (in seconds) between requests to avoid overloading the BDTD server."
    )
    parser.add_argument(
        '--page-size', 
        type=int, 
        default=100, 
        help="Number of records to fetch per pagination request (max 100)."
    )
    parser.add_argument(
        '--keywords',
        type=str,
        nargs='+',
        default=[
            "desenvolvimento regional", 
            "políticas públicas", 
            "planejamento urbano"
        ],
        help="List of keywords/phrases to search for."
    )
    parser.add_argument(
        '--export-excel',
        type=str,
        default='bdtd_metadata.xlsx',
        help="Path where the final sheet should be saved (supports .xlsx, .csv, .json)."
    )
    parser.add_argument(
        '--config',
        type=str,
        default=None,
        help="Path to a JSON or Excel configuration file. If omitted, checks for bdtd_config.json first, then bdtd_config.xlsx."
    )
    parser.add_argument(
        '--type',
        type=str,
        default='AllFields',
        help="Search field type (AllFields, Title, Author, Subject, Advisor)."
    )
    parser.add_argument(
        '--sort',
        type=str,
        default='year',
        help="Sorting order (year, year asc, relevance, title, author)."
    )
    parser.add_argument(
        '--filter',
        type=str,
        nargs='*',
        default=[],
        help="List of Solr filters (e.g. format:masterThesis institution:USP)."
    )
    parser.add_argument(
        '--fast-mode',
        action='store_true',
        help="Skip individual webpage scraping (scrape_details=False) for maximum speed."
    )
    
    args = parser.parse_args()  # Processa e lê os argumentos passados no terminal
    setup_logging()  # Configura logging limpo no console
    
    config_file: Optional[str] = args.config  # Obtém o caminho do arquivo de configuração
    use_config: bool = False  # Flag indicativa de uso de arquivo de configuração
    config_type: Optional[str] = None  # Tipo do arquivo (json ou excel)
    
    if config_file:  # Se o usuário especificou um arquivo via parâmetro --config
        if os.path.exists(config_file):  # Se o arquivo existir no disco
            use_config = True  # Ativa uso de configuração
            config_type = 'json' if config_file.lower().endswith('.json') else 'excel'  # Determina o tipo pela extensão
    else:
        # Se não informou --config, verifica se existem os arquivos padrão no diretório atual
        if os.path.exists('bdtd_config.json'):
            config_file = 'bdtd_config.json'
            use_config = True
            config_type = 'json'
        elif os.path.exists('bdtd_config.xlsx'):
            config_file = 'bdtd_config.xlsx'
            use_config = True
            config_type = 'excel'
            
    scrape_details: bool = not args.fast_mode  # Ativa raspagem de detalhes se não estiver em fast-mode
    db_path: str
    export_excel: str
    limit: Optional[int]
    delay: float
    target_keywords: List[str]
    search_type: str
    sort_order: str
    filters: List[str]

    if use_config and config_file:  # Se houver arquivo de configuração identificado
        try:
            config: Dict[str, Any]
            if config_type == 'json':
                config = read_json_config_file(config_file)  # Lê o arquivo JSON
            else:
                config = read_config_file(config_file)  # Lê o arquivo Excel
                
            # Extrai os parâmetros lidos do arquivo de configuração
            db_path = config['db_path']
            export_excel = config['export_excel']
            limit = config['limit']
            delay = config['delay']
            target_keywords = config['keywords']
            search_type = config.get('search_type', 'AllFields')
            sort_order = config.get('sort_order', 'year')
            filters = config.get('filters', [])
            if 'scrape_details' in config and not args.fast_mode:
                scrape_details = config['scrape_details']
            logger.info(f"Loaded config from {config_file}: DB={db_path}, Export={export_excel}, Limit={limit}, Delay={delay}, Type={search_type}, Sort={sort_order}, Filters={filters}, ScrapeDetails={scrape_details}, Keywords={target_keywords}")
        except Exception as e:  # Se houver erro ao carregar o arquivo de configuração
            logger.error(f"Failed to read configuration file: {e}. Falling back to CLI arguments.")
            # Restaura para os parâmetros fornecidos diretamente na linha de comando
            db_path = args.db_path
            export_excel = args.export_excel
            limit = args.limit
            delay = args.delay
            target_keywords = args.keywords
            search_type = args.type
            sort_order = args.sort
            filters = args.filter
    else:
        # Se nenhum arquivo de configuração existia, gera os modelos padrões para o usuário
        if not config_file:
            try:
                if not os.path.exists('bdtd_config.json'):
                    create_json_config_template('bdtd_config.json')  # Cria o bdtd_config.json padrão
                if not os.path.exists('bdtd_config.xlsx'):
                    create_config_template('bdtd_config.xlsx')  # Cria o bdtd_config.xlsx padrão
            except Exception as e:
                logger.warning(f"Could not create configuration templates: {e}")
                
        # Utiliza os argumentos vindos da linha de comando
        db_path = args.db_path
        export_excel = args.export_excel
        limit = args.limit
        delay = args.delay
        target_keywords = args.keywords
        search_type = args.type
        sort_order = args.sort
        filters = args.filter

    if not target_keywords:  # Se nenhuma palavra-chave foi definida
        logger.error("No keywords specified. Please define keywords via CLI or config file.")
        return  # Interrompe a execução

    try:
        # Executa a função principal de execução do pipeline
        success: bool = run_harvest(
            keywords=target_keywords,
            db_path=db_path,
            export_path=export_excel,
            limit=limit,
            delay=delay,
            page_size=args.page_size,
            search_type=search_type,
            sort_order=sort_order,
            filters=filters,
            scrape_details=scrape_details
        )
        if success:
            logger.info("Pipeline executed successfully.")  # Registra conclusão com sucesso
        else:
            logger.error("Pipeline finished with errors.")  # Registra finalização com erro
    except KeyboardInterrupt:  # Se o usuário pressionou Ctrl+C
        logger.warning("\nPipeline interrupted by user.")  # Registra interrupção pelo usuário
    except Exception as e:  # Se ocorreu um erro não tratado
        logger.critical(f"Pipeline crashed due to an unhandled exception: {e}", exc_info=True)  # Registra falha crítica


# Ponto de entrada do script quando executado diretamente
if __name__ == '__main__':
    main()  # Executa a função main
